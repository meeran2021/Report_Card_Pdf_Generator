from reportlab.pdfgen import canvas
from PyPDF2 import PdfFileWriter, PdfFileReader
import pandas as pd
import openpyxl
import io
location=r'A:\Learning Materials\Codings\VS Code\others\Wis\New folder'
xlFile= r'\Dummy Data.xlsx'
pdfName= r'\Report Card1.pdf'

logo=r'\Logo.jpg'
xlsxPath = location + xlFile


def createLayout(can,r):
    for _ in range(10):
        can.drawString(200, 800, "Joint Entrance Examination")
        can.drawString(260, 585, "Answer Key")

    #Round
    cellVal = sheet.cell(row = 2, column = 2)
    can.drawString(240, 780, str(cellVal.value)+" :-")
    x=30
    y=755
    for col in colList:
        if col==20:
            y=110
        cellVal = sheet.cell(row = 2, column = col)
        can.drawString(x,y, str(cellVal.value)+" :-")
        y-=15
    x=70    #
    y=z=560
    for col in range(14,20):
        cellVal = sheet.cell(row = 2, column = col)
        st=str(cellVal.value).split()  
        for val in range(len(st)):
            for _ in range(3):
                can.drawString(x, y, st[val])
            y-=15
        if col==17:
            x+=65    
        y=z
        x+=60
    printData(can,r)
    printResult(can,r)
    can.showPage()
    can.save()


def printData(can,r):
    r=r*25
    
    y=755
    dict1={12:155  , 3:110 , 4:110 , 9:85 , 10:120 , 7:80 , 8:135 , 11:140 , 13:160}    #, 6:155
    for val in dict1:
        cellVal = sheet.cell(row = r+3, column = val)
        can.drawString(dict1[val], y, str(cellVal.value))
        if val==12 or val==4:
            y-=15
        y-=15
    
    can.drawString(110, 695, st.iloc[r+3,0])
    can.drawString(155, 740, str(regNo.iloc[r+3,0]))
    cellVal = sheet.cell(row = r+3, column = 2)
    can.drawString(290, 780, str(cellVal.value))


def printResult(can,r):
    r=r*25
    x=80    #
    y=z=510
    for col in range(14,20):
        for a in range(r,r+25):
            if col==17:
                can.drawString(x,y,outcomes.iloc[a,0])
            elif col==19:
                can.drawString(x,y,str(score.iloc[a,0]))
            else:
                cellVal = sheet.cell(row = a+3, column = col)
                can.drawString(x, y, str(cellVal.value))
            y-=15
        if col==17 or col==19:
            x+=60
        x+=60
        y=z
    for _ in range(5):
        cellVal = sheet.cell(row = r+3, column = 20)
        can.drawString(105,110, str(cellVal.value))

def insertPic(imgFile, newPdf,finalPdf):
    
    logoFile= location + logo

    instance = io.BytesIO()
    can = canvas.Canvas(instance)

    can.drawImage(logoFile, 20, 720, width=60, preserveAspectRatio=True, mask='auto')
    can.drawImage(imgFile, 420, 650, width=120,height=120, preserveAspectRatio=True, mask='auto')
    can.showPage()
    can.save()
 
    instance.seek(0) 
    newPdf1 = PdfFileReader(instance)
    existingPdf = PdfFileReader(open(newPdf, "rb"))
    output = PdfFileWriter()
 
    for i in range(len(existingPdf.pages)):
        page = existingPdf.getPage(i)
        page.mergePage(newPdf1.getPage(i))
        output.addPage(page)
 
    outputStream = open(finalPdf, "wb")
    output.write(outputStream)
    outputStream.close()


if __name__ == '__main__':
    wb = openpyxl.load_workbook(xlsxPath)
    sheet = wb.active

    colList=[12,6,3,4,5,9,10,7,8,11,13,20]
    # xList=[30,30,30,30,30,30,30,30,30,30,30,30]
    yList=[755,740,725,710,695,680,665,650,635,620,605,110]

    st= pd.read_excel(xlsxPath,usecols='E', dtype=str, skiprows=1)
    regNo= pd.read_excel(xlsxPath,usecols='F', dtype=str, skiprows=1)
    outcomes= pd.read_excel(xlsxPath,usecols='Q', skiprows=1)
    score= pd.read_excel(xlsxPath,usecols='S', skiprows=1)

    pdfList= ['\ABC1.pdf','\ABC2.pdf','\ABC3.pdf','\ABC4.pdf','\ABC5.pdf']
    resultPdfList= ['\Result ABC1.pdf','\Result ABC2.pdf','\Result ABC3.pdf','\Result ABC4.pdf','\Result ABC5.pdf']
    imgList = ['\ABC1 XYZ1.png','\ABC2 XYZ2.png','\ABC3 XYZ3.png','\ABC4 XYZ4.png','\ABC5 XYZ5.png']
    for i in range(len(pdfList)):
        newPdf = location + pdfList[i]
        finalPdf = location + resultPdfList[i]
        imgFile = location + imgList[i]
        can = canvas.Canvas(newPdf)
        createLayout(can,i)
        insertPic(imgFile,newPdf,finalPdf)